import argparse
import os
import openpyxl

def load_text(filename):
    with open(filename, 'r', encoding='utf-8') as f:
        text = f.read()
    return text

def generate_ngrams(text, n):
    ngrams = []
    for i in range(len(text)-n+1):
        ngrams.append(text[i:i+n])
    return ngrams

def compare_texts(text1, text2, n):
    ngrams1 = generate_ngrams(text1, n)
    ngrams2 = generate_ngrams(text2, n)
    set1 = set(ngrams1)
    set2 = set(ngrams2)
    overlap = set1 & set2
    score = len(overlap) / max(len(set1), len(set2))
    return score

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Compare two texts using n-gram similarity')
    parser.add_argument('mainfile', type=str, help='filename of the main text')
    parser.add_argument('dir', type=str, help='directory containing the other texts')
    parser.add_argument('--n', type=int, default=2, help='value of n for the n-grams')
    parser.add_argument('--output', type=str, default='output.xlsx', help='output Excel filename')
    args = parser.parse_args()

    text1 = load_text(args.mainfile)

    # iterate over files in the directory
    scores = []
    for filename in os.listdir(args.dir):
        if filename.endswith('.txt'):
            filepath = os.path.join(args.dir, filename)
            text2 = load_text(filepath)
            similarity = compare_texts(text1, text2, args.n)
            scores.append((filename, similarity))

    # sort by descending similarity
    scores.sort(key=lambda x: x[1], reverse=True)

    # save results to Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Similarities'
    ws.cell(row=1, column=1, value='Filename')
    ws.cell(row=1, column=2, value='Similarity')
    for i, (filename, similarity) in enumerate(scores):
        ws.cell(row=i+2, column=1, value=filename)
        ws.cell(row=i+2, column=2, value=similarity)
    wb.save(args.output)

    print(f'Results saved to {args.output}')
